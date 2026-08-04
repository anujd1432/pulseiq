import os
from datetime import datetime
from app.core.celery_app import celery_app
from app.core.database import SessionLocal
from app.routers.reports import gather_kpi_data
from app.core.email_utils import send_email_with_attachment

REPORTS_DIR = "/app/generated_reports"


def _build_narrative(data: dict) -> str:
    icu_rate = data.get("ICU Occupancy Rate (%)", 0)
    low_stock = data.get("Low Stock Medicine Alerts", 0)
    active_admissions = data.get("Active Admissions", 0)

    lines = []
    if icu_rate >= 80:
        lines.append(f"ICU occupancy is high at {icu_rate}%, warranting close monitoring.")
    elif icu_rate >= 50:
        lines.append(f"ICU occupancy is at a moderate {icu_rate}%.")
    else:
        lines.append(f"ICU occupancy is comfortable at {icu_rate}%.")

    if low_stock > 0:
        lines.append(f"{low_stock} medicine(s) are at or below reorder level and need procurement attention.")
    else:
        lines.append("Medicine stock levels are healthy across the board.")

    lines.append(f"There are currently {active_admissions} active admissions.")
    return " ".join(lines)


@celery_app.task(name="app.tasks.report_tasks.generate_scheduled_report")
def generate_scheduled_report(period: str = "weekly"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    db = SessionLocal()
    try:
        data = gather_kpi_data(db)
    finally:
        db.close()

    os.makedirs(REPORTS_DIR, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"pulseiq_{period}_report_{timestamp}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = f"PulseIQ {period.capitalize()} Report"

    ws["A1"] = f"PulseIQ {period.capitalize()} KPI Report"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"

    narrative = _build_narrative(data)
    ws["A3"] = "Summary:"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = narrative
    ws.merge_cells("A4:B4")
    ws["A4"].alignment = ws["A4"].alignment.copy(wrap_text=True)

    ws["A6"] = "Metric"
    ws["B6"] = "Value"
    ws["A6"].font = Font(bold=True)
    ws["B6"].font = Font(bold=True)
    ws["A6"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws["B6"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    row = 7
    for key, value in data.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    wb.save(filepath)

    email_result = send_email_with_attachment(filepath, period)

    return {"status": "success", "file_generated": filepath, "period": period, "email": email_result}
