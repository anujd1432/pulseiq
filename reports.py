from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime, timedelta
from io import BytesIO
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.patient import Patient
from app.models.admission import Admission
from app.models.bed import Bed
from app.models.doctor import Doctor
from app.models.ambulance import Ambulance
from app.models.inventory import Inventory
from app.models.medicine import Medicine

router = APIRouter(prefix="/reports", tags=["Reports"])


def gather_kpi_data(db: Session):
    total_patients = db.query(func.count(Patient.id)).scalar()
    active_admissions = db.query(func.count(Admission.id)).filter(Admission.status == "admitted").scalar()

    total_icu_beds = db.query(func.count(Bed.id)).filter(Bed.bed_type == "icu").scalar()
    occupied_icu_beds = db.query(func.count(Bed.id)).filter(Bed.bed_type == "icu", Bed.status == "occupied").scalar()
    icu_occupancy_rate = round((occupied_icu_beds / total_icu_beds) * 100, 2) if total_icu_beds > 0 else 0

    total_beds = db.query(func.count(Bed.id)).scalar()
    available_beds = db.query(func.count(Bed.id)).filter(Bed.status == "available").scalar()

    discharged = db.query(Admission).filter(Admission.discharge_date.isnot(None)).all()
    alos = round(sum((a.discharge_date - a.admission_date).days for a in discharged) / len(discharged), 2) if discharged else 0

    total_doctors = db.query(func.count(Doctor.id)).scalar()

    total_ambulances = db.query(func.count(Ambulance.id)).scalar()
    available_ambulances = db.query(func.count(Ambulance.id)).filter(Ambulance.status == "available").scalar()

    low_stock = db.query(Inventory).filter(Inventory.current_stock <= Inventory.reorder_level).count()

    return {
        "Total Patients": total_patients,
        "Active Admissions": active_admissions,
        "ICU Occupancy Rate (%)": icu_occupancy_rate,
        "Total Beds": total_beds,
        "Available Beds": available_beds,
        "Average Length of Stay (days)": alos,
        "Total Doctors": total_doctors,
        "Total Ambulances": total_ambulances,
        "Available Ambulances": available_ambulances,
        "Low Stock Medicine Alerts": low_stock,
    }


@router.get("/excel")
def export_excel_report(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    data = gather_kpi_data(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "PulseIQ KPI Report"

    ws["A1"] = "PulseIQ Hospital KPI Report"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"

    ws["A4"] = "Metric"
    ws["B4"] = "Value"
    ws["A4"].font = Font(bold=True)
    ws["B4"].font = Font(bold=True)
    ws["A4"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws["B4"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    row = 5
    for key, value in data.items():
        ws[f"A{row}"] = key
        ws[f"B{row}"] = value
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pulseiq_kpi_report.xlsx"}
    )


@router.get("/pdf")
def export_pdf_report(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    data = gather_kpi_data(db)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("PulseIQ Hospital KPI Report", styles["Title"]))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    elements.append(Spacer(1, 0.5 * cm))

    table_data = [["Metric", "Value"]] + [[k, str(v)] for k, v in data.items()]
    table = Table(table_data, colWidths=[10 * cm, 5 * cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=pulseiq_kpi_report.pdf"}
    )
